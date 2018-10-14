from prog import bsqlproc
import openpyxl as pxl


def model_export(datatab, opt, model=None, view=None):
    """ model is not None => colors and fonts
        view is not None => cell sizes
    """
    if opt.format == 'plain text':
        return plain_text_export(datatab, opt)
    elif opt.format == 'xlsx':
        return xlsx_export(datatab, opt, model, view)
    else:
        raise NotImplementedError


def _get_data_lines(datatab, opt):
    lines = []
    if opt.with_caption:
        lines.append([datatab.column_caption(i)
                      for i in range(datatab.n_cols())])

    for i in range(datatab.n_rows()):
        rvals = [None]*datatab.n_cols()
        for j, col in enumerate(datatab.visible_columns):
            rvals[j] = datatab.tab.get_value(i, j)

            if rvals[j] is not None and not opt.numeric_enums:
                rvals[j] = col.repr(rvals[j])

            if rvals[j] is None and col.is_category() and\
                    datatab.n_subrows(i) > 1:
                if opt.grouped_categories == 'Comma separated':
                    vals = datatab.get_raw_subvalues(i, j)
                    if not opt.numeric_enums:
                        vals = map(col.repr, vals)
                    rvals[j] = ','.join(map(str, vals))
                elif opt.grouped_categories == 'Unique count':
                    n_uni = datatab.n_subdata_unique(i, j)
                    rvals[j] = bsqlproc.group_repr(n_uni)
                elif opt.grouped_categories == 'None':
                    rvals[j] = ''

        lines.append(rvals)

    if not opt.with_id:
        lines = [x[1:] for x in lines]

    return lines


def plain_text_export(datatab, opt):
    lines = _get_data_lines(datatab, opt)

    if not opt.with_id:
        lines = [x[1:] for x in lines]
    lines = ['\t'.join(map(str, x)) + '\n' for x in lines]

    with open(opt.filename, 'w') as fid:
        fid.writelines(lines)
    fid.close()


def xlsx_export(datatab, opt, model, view):
    lines = _get_data_lines(datatab, opt)

    wb = pxl.Workbook()
    ws1 = wb.active
    ws1.title = datatab.table_name()
    for line in lines:
        ws1.append(list(line))

    if opt.with_formatting and model is not None:
        from bgui import tmodel
        from PyQt5 import QtCore
        for i, row in enumerate(ws1.rows):
            for j, cell in enumerate(row):
                ir, jr = i + 2, j + 1
                if opt.with_caption:
                    ir -= 1
                if opt.with_id:
                    jr -= 1
                index = model.createIndex(ir, jr)
                ft = model.data(index, QtCore.Qt.FontRole)
                if model.use_coloring() or ir < 2 or jr < 1:
                    # 1) colors
                    fg, bg = model.data(index, tmodel.TabModel.ColorsRole)
                    clr = pxl.styles.colors.Color(rgb=bg.name()[1:])
                    my_fill = pxl.styles.fills.PatternFill(
                            patternType='solid', fgColor=clr)
                    cell.fill = my_fill
                    # 2) fonts
                    pft = pxl.styles.Font(
                            sz=ft.pointSize(), color=fg.name()[1:],
                            italic=ft.italic(), bold=ft.bold())
                    cell.font = pft
                else:
                    pft = pxl.styles.Font(sz=ft.pointSize())
                    cell.font = pft

        # 3) cell sizes
        if view is not None:
            for i, col in enumerate(ws1.iter_cols(max_row=1)):
                n = col[0].column
                w = view.horizontalHeader().sectionSize(i)
                ws1.column_dimensions[n].width = w / 5.54
            h = view.verticalHeader().sectionSize(1)
            for i, row in enumerate(ws1.iter_rows(max_col=1)):
                ws1.row_dimensions[i+1].height = h

    wb.save(opt.filename)


def qmodel_xlsx_export(model, fname, hheader=False, vheader=False):
    from PyQt5 import QtCore
    wb = pxl.Workbook()
    ws1 = wb.active
    nrows = model.rowCount()
    ncols = model.columnCount()
    if hheader:
        vals = [model.headerData(i, QtCore.Qt.Horizontal,
                                 QtCore.Qt.DisplayRole) for i in range(nrows)]
        if vheader:
            vals.insert(0, '')
        ws1.append(vals)
    for i in range(nrows):
        vals = [model.data(model.index(i, j), QtCore.Qt.DisplayRole)
                for j in range(ncols)]
        if vheader:
            vals.insert(0, model.headerData(i, QtCore.Qt.Vertical,
                                            QtCore.Qt.DisplayRole))
        ws1.append(vals)
    wb.save(fname)


def get_unused_tmp_file(ext):
    import tempfile
    if ext:
        s = '.' + ext
    else:
        s = ''
    handle, fn = tempfile.mkstemp(suffix=s)
    return fn
