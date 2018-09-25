import openpyxl as pxl
from prog import comproj
from bdata import bcol
from bdata import dtab


def split_plain_text(fname, options):
    """ used options attributes:
            firstline, lastline, comment_sign, ignore_blank, row_sep, col_sep
            colcount.
        returns equal column size 2d array of stripped text entries
    """
    with open(fname, 'r') as f:
        lines = f.readlines()
    f.close()

    # firstline-lastline
    firstline = max(0, options.firstline - 1)
    lastline = options.lastline if options.lastline > 0 else len(lines)-1
    lines = lines[firstline:lastline+1]
    # remove text after comments
    if options.comment_sign:
        for i, line in enumerate(lines):
            pos = line.find(options.comment_sign)
            if pos >= 0:
                lines[i] = line[:pos]
    # remove blank lines
    if options.ignore_blank:
        lines = list(filter(lambda x: len(x.strip()) > 0, lines))

    # reassemble lines if row separator is not a new line
    if options.row_sep != "newline":
        lines = " ".join(lines)
        if options.row_sep != "no (use column count)":
            lines = lines.split(options.row_sep)
        else:
            lines = [lines]

    # assemble columns
    if options.col_sep == 'whitespaces':
        for i, line in enumerate(lines):
            lines[i] = line.split()
    elif options.col_sep == 'tabular':
        for i, line in enumerate(lines):
            lines[i] = line.split('\t')
    elif options.col_sep == 'in double quotes':
        for i, line in enumerate(lines):
            lines[i] = line.split('"')[1::2]
    else:
        for i, line in enumerate(lines):
            lines[i] = line.split(options.col_sep)

    if len(lines) == 0:
        raise Exception("No data were loaded")

    # row split by column count
    if options.row_sep == "no (use column count)":
        lines2 = [lines[0][x:x+options.colcount]
                  for x in range(0, len(lines[0]), options.colcount)]
        lines = lines2

    # equal column count
    if options.colcount <= 0:
        cc = max([len(x) for x in lines])
    else:
        cc = options.colcount
    for i, line in enumerate(lines):
        if len(line) > cc:
            lines[i] = line[:cc]
        elif len(line) < cc:
            lines[i] = line + [""] * (cc - len(line))

    # final strip
    for line in lines:
        for j, x in enumerate(line):
            line[j] = x.strip()
    return lines


def read_xlsx_sheets(fname):
    return pxl.load_workbook(fname, read_only=True, data_only=True).sheetnames


def parse_xlsx_file(fname, options):
    doc = pxl.load_workbook(fname, read_only=True, data_only=True)
    sh = doc[options.sheetname]
    if options.range != '':
        min_col, min_row, max_col, max_row = pxl.utils.range_boundaries(
                options.range.upper())
    else:
        min_col, min_row = sh.min_column, sh.min_row
        max_col, max_row = sh.max_column, sh.max_row
    ret = []

    for row in sh.iter_rows(None, min_row, max_row, min_col, max_col):
        ret.append(list(map(lambda x: str(x.value)
                            if x.value is not None else '', row)))

    return ret


def autodetect_types(tab, maxsize=10):
    from ast import literal_eval
    if len(tab) == 0:
        return []

    nrows = len(tab)
    ncols = max([len(x) for x in tab])

    ret = []
    for icol in range(ncols):
        r = "INT"  # 3-INT, 2-REAL, 1-TEXT
        s = 0
        for irow in range(nrows):
            v = tab[irow][icol]
            if v is not "" and v is not None:
                try:
                    x = literal_eval(v)
                    if type(x) is float:
                        r = "REAL"
                    elif type(x) is not int:
                        raise
                    s += 1
                    if s >= maxsize:
                        break
                except:
                    r = "TEXT"
                    break
        ret.append(r)
    return ret


def explicit_table(tab_name, colformats, tab, proj):
    """ Assembles a table from given python data.
        colformats = [(name, dt_type, dict_name), ...]
        tab[][] - table in python types format. None for NULL.

        !!! This procedure modifies tab (converts string values to format
        values). Use deepcopy in order to keep tab unchanged.
    """
    for i, t in enumerate(tab):
        t.insert(0, i+1)

    def init_columns(self):
        self.all_columns = []
        self.all_columns.append(bcol.build_id())
        for name, dt_type, dict_name in colformats:
            col = bcol.explicit_build(self.proj, name, dt_type, dict_name)
            self.all_columns.append(col)

    def fill_ttab(self):
        cols, sqlcols = [], []
        for c in self.all_columns:
            cols.append(c)
            sqlcols.append(c.sql_line())

        for i in range(len(tab)):
            for j in range(len(tab[i])):
                tab[i][j] = cols[j].from_repr(tab[i][j])

        pholders = ','.join(['?'] * (len(self.all_columns)))
        qr = 'INSERT INTO "{tabname}" ({collist}) VALUES ({ph})'.format(
                tabname=self.ttab_name,
                collist=", ".join(sqlcols),
                ph=pholders)
        self.query(qr, tab)

    return dtab.DataTable(tab_name, proj, init_columns, fill_ttab, True)


class _ImportTab(comproj.NewTabCommand):
    def __init__(self, proj, fname, opt, delegate):
        super().__init__(proj)
        self.fname = fname
        self.opt = opt

        self.delegate = delegate
        self.tab = None
        self.tps = None
        self.dnames = None
        self.caps = None

    def _prebuild(self):
        self.tab = self.delegate(self.fname, self.opt)
        if self.opt.read_cap:
            self.caps = self.tab[0][:]
            self.tab = self.tab[1:]
        else:
            self.caps = ["Column {}".format(i+1)
                         for i in range(len(self.tab[0]))]
        self.tps = autodetect_types(self.tab)
        self.dnames = [None] * len(self.tps)

    def _clear(self):
        self.tab = None
        self.caps = None
        self.tps = None
        self.dnames = None
        super()._clear()

    def _get_table(self):
        if self.tab is None:
            self._prebuild()
        a = [(c, tp, d) for c, tp, d in zip(self.caps, self.tps, self.dnames)]
        return explicit_table(self.opt.tabname, a, self.tab, self.proj)


class ImportTabFromTxt(_ImportTab):
    def __init__(self, proj, fname, opt):
        """ opt.firstline = 0
            opt.lastline = -1
            opt.read_cap = True
            opt.comment_sign = '#'
            opt.ignore_blank = True
            opt.col_sep = 'whitespaces'
            opt.row_sep = 'newline'
            opt.colcount = -1
            opt.tabname = 't1'
        """
        super().__init__(proj, fname, opt, split_plain_text)


class ImportTabFromXlsx(_ImportTab):
    def __init__(self, proj, fname, opt):
        """ opt.sheetname = 'tab1'
            opt.range = ''
            opt.read_cap = True
            opt.tabname = 't1'
        """
        super().__init__(proj, fname, opt, parse_xlsx_file)
